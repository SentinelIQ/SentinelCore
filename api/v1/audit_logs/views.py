from rest_framework import viewsets, generics, filters
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework import status
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from drf_spectacular.utils import extend_schema, OpenApiParameter, extend_schema_view
from auditlog.models import LogEntry
from .serializers import AuditLogSerializer, AuditLogListSerializer
from .filters import AuditLogFilter
from api.core.responses import StandardResponse, success_response
from api.core.pagination import StandardResultsSetPagination
from api.v1.audit_logs.enums import EntityTypeEnum, ActionTypeEnum
from api.core.utils.export import ExportMixin
from api.core.rbac import HasEntityPermission
from django.http import HttpResponse
import csv
import json


@extend_schema_view(
    list=extend_schema(
        summary="Listar logs de auditoria",
        description="Retorna uma lista paginada de logs de auditoria com opções de filtro",
        tags=["Authentication & Access Control"],
        parameters=[
            OpenApiParameter(name="entity_type", description="Filtrar por tipo de entidade", enum=EntityTypeEnum, required=False),
            OpenApiParameter(name="action", description="Filtrar por ação", enum=ActionTypeEnum, required=False),
            OpenApiParameter(name="entity_id", description="Filtrar por ID da entidade", type=str, required=False),
            OpenApiParameter(name="username", description="Filtrar por nome de usuário", type=str, required=False),
            OpenApiParameter(name="company_id", description="Filtrar por ID da empresa", type=str, required=False),
            OpenApiParameter(name="date_from", description="Filtrar por intervalo de data (de), formato: YYYY-MM-DD HH:MM:SS", type=str, required=False),
            OpenApiParameter(name="date_to", description="Filtrar por intervalo de data (até), formato: YYYY-MM-DD HH:MM:SS", type=str, required=False),
            OpenApiParameter(name="period", description="Filtrar por período predefinido (today, yesterday, week, month, year)", type=str, required=False),
            OpenApiParameter(name="search", description="Pesquisar em vários campos", type=str, required=False),
        ]
    ),
    retrieve=extend_schema(
        summary="Detalhar log de auditoria",
        description="Retorna informações detalhadas sobre um log de auditoria específico",
        tags=["Authentication & Access Control"]
    ),
    export=extend_schema(
        summary="Exportar logs de auditoria",
        description="Exportar logs de auditoria filtrados para CSV, JSON ou Excel",
        tags=["Authentication & Access Control"],
        parameters=[
            OpenApiParameter(name="format", description="Formato de exportação (csv, json, excel)", type=str, required=False, default="csv"),
        ]
    )
)
class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint para logs de auditoria.
    
    Este viewset fornece operações de listagem, detalhamento e exportação de logs de auditoria.
    Suporta filtragem por tipo de entidade, ação, intervalo de data e outros campos.
    
    Os logs de auditoria fornecem um histórico completo de todas as ações realizadas no sistema,
    incluindo quem realizou a ação, quando foi realizada e o que foi afetado.
    """
    queryset = LogEntry.objects.all()
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = AuditLogFilter
    search_fields = ['actor__username', 'object_repr', 'additional_data', 'object_pk', 'action']
    ordering_fields = ['timestamp', 'action', 'actor__username', 'object_repr', 'remote_addr']
    ordering = ['-timestamp']
    
    def get_serializer_class(self):
        """
        Retornar serializer apropriado:
        - AuditLogListSerializer para listagem
        - AuditLogSerializer para detalhamento
        """
        if self.action == 'list':
            return AuditLogListSerializer
        return AuditLogSerializer
    
    def get_queryset(self):
        """
        Limitar queryset à empresa do usuário atual, a menos que seja superusuário
        """
        queryset = super().get_queryset()
        
        user = self.request.user
        if not user.is_superuser:
            # Usuários comuns só podem ver logs de sua empresa
            if hasattr(user, 'company') and user.company:
                company_id = str(user.company.id)
                # Filtrar logs pelo company_id nos dados adicionais
                queryset = queryset.filter(additional_data__company_id=company_id)
            else:
                # Se o usuário não tiver empresa, mostrar apenas seus próprios logs
                queryset = queryset.filter(actor=user)
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """
        Exportar logs de auditoria para CSV, JSON ou Excel.
        
        Esta action aplica os mesmos filtros da action list,
        mas retorna um arquivo em vez de uma resposta API paginada.
        
        Formatos suportados:
        - csv: arquivo CSV (padrão)
        - json: arquivo JSON
        - excel: arquivo Excel
        """
        # Obter queryset filtrado sem paginação
        queryset = self.filter_queryset(self.get_queryset())
        
        # Obter formato de exportação
        export_format = request.query_params.get('format', 'csv').lower()
        
        if export_format == 'json':
            return self._export_json(queryset)
        elif export_format == 'excel':
            return self._export_excel(queryset)
        else:
            # Padrão para CSV
            return self._export_csv(queryset)
    
    def _export_csv(self, queryset):
        """Exportar logs de auditoria para formato CSV."""
        import csv
        from django.http import HttpResponse
        
        # Criar resposta CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="audit_logs.csv"'
        
        # Escrever cabeçalho CSV
        writer = csv.writer(response)
        writer.writerow([
            'Timestamp', 'User', 'Action', 'Entity Type', 
            'Entity ID', 'Entity Name', 'Company', 'Status',
            'IP Address', 'Request Method', 'Request Path'
        ])
        
        # Escrever linhas de dados
        for log in queryset:
            # Extrair dados do LogEntry e additional_data
            additional_data = log.additional_data or {}
            entity_type = additional_data.get('entity_type', '')
            entity_id = log.object_pk
            entity_name = log.object_repr
            status = additional_data.get('response_status', '')
            company_name = additional_data.get('company_name', '')
            request_method = additional_data.get('request_method', '')
            request_path = additional_data.get('request_path', '')
            username = log.actor.username if log.actor else 'System'
            
            writer.writerow([
                log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                username,
                log.get_action_display(),
                entity_type,
                entity_id,
                entity_name,
                company_name,
                status,
                log.remote_addr,
                request_method,
                request_path
            ])
        
        return response
    
    def _export_json(self, queryset):
        """Exportar logs de auditoria para formato JSON."""
        import json
        from django.http import HttpResponse
        
        # Converter queryset para lista de dicts
        logs_data = []
        for log in queryset:
            # Extrair dados do LogEntry e additional_data
            additional_data = log.additional_data or {}
            entity_type = additional_data.get('entity_type', '')
            entity_id = log.object_pk
            entity_name = log.object_repr
            status = additional_data.get('response_status', '')
            company_name = additional_data.get('company_name', '')
            request_method = additional_data.get('request_method', '')
            request_path = additional_data.get('request_path', '')
            
            logs_data.append({
                'id': str(log.id),
                'timestamp': log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                'user': log.actor.username if log.actor else 'System',
                'action': log.get_action_display(),
                'entity_type': entity_type,
                'entity_id': entity_id,
                'entity_name': entity_name,
                'company': company_name,
                'status': status,
                'ip_address': log.remote_addr,
                'request_method': request_method,
                'request_path': request_path
            })
        
        # Criar resposta JSON
        response = HttpResponse(json.dumps(logs_data, indent=2), content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename="audit_logs.json"'
        
        return response
    
    def _export_excel(self, queryset):
        """Exportar logs de auditoria para formato Excel."""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from io import BytesIO
            from django.http import HttpResponse
            
            # Criar workbook e sheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Audit Logs"
            
            # Escrever cabeçalhos
            headers = [
                'Timestamp', 'User', 'Action', 'Entity Type', 
                'Entity ID', 'Entity Name', 'Company', 'Status',
                'IP Address', 'Request Method', 'Request Path'
            ]
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = openpyxl.styles.Font(bold=True)
            
            # Escrever dados
            for row_idx, log in enumerate(queryset, 2):
                # Extrair dados do LogEntry e additional_data
                additional_data = log.additional_data or {}
                entity_type = additional_data.get('entity_type', '')
                entity_id = log.object_pk
                entity_name = log.object_repr
                status = additional_data.get('response_status', '')
                company_name = additional_data.get('company_name', '')
                request_method = additional_data.get('request_method', '')
                request_path = additional_data.get('request_path', '')
                
                cols = [
                    log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                    log.actor.username if log.actor else 'System',
                    log.get_action_display(),
                    entity_type,
                    entity_id,
                    entity_name,
                    company_name,
                    status,
                    log.remote_addr or '',
                    request_method,
                    request_path
                ]
                
                for col_idx, value in enumerate(cols, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    
            # Auto-size colunas
            for col_idx in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].auto_size = True
            
            # Salvar workbook no buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Criar resposta
            response = HttpResponse(
                buffer.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="audit_logs.xlsx"'
            
            return response
        except ImportError:
            # Fallback para CSV se openpyxl não estiver disponível
            return self._export_csv(queryset)
    
    def list(self, request, *args, **kwargs):
        """Listar todos os logs de auditoria com filtragem"""
        queryset = self.filter_queryset(self.get_queryset())
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
            
        serializer = self.get_serializer(queryset, many=True)
        return success_response(data=serializer.data)
        
    def retrieve(self, request, *args, **kwargs):
        """Recuperar um log de auditoria específico"""
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return success_response(
            data=serializer.data,
            message=f"Log de auditoria {instance.id} recuperado com sucesso"
        )


class AuditLogList(generics.ListAPIView):
    """
    Listar todos os logs de auditoria com filtros.
    
    Suporta filtragem por ação, tipo de entidade, data e usuário.
    """
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated, HasEntityPermission]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = AuditLogFilter
    ordering_fields = ['timestamp', 'action', 'object_repr']
    ordering = ['-timestamp']
    
    # Adicionar entity_type para verificação de permissão RBAC
    entity_type = 'audit_log'

    def get_queryset(self):
        """
        Limitar queryset à empresa do usuário atual, a menos que seja superusuário
        """
        queryset = LogEntry.objects.all()
        
        user = self.request.user
        if not user.is_superuser:
            # Usuários comuns só podem ver logs de sua empresa
            if hasattr(user, 'company') and user.company:
                company_id = str(user.company.id)
                # Filtrar logs pelo company_id nos dados adicionais
                queryset = queryset.filter(additional_data__company_id=company_id)
            else:
                # Se o usuário não tiver empresa, mostrar apenas seus próprios logs
                queryset = queryset.filter(actor=user)
        
        return queryset
    
    def list(self, request, *args, **kwargs):
        """Listar todos os logs de auditoria com filtragem"""
        queryset = self.filter_queryset(self.get_queryset())
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
            
        serializer = self.get_serializer(queryset, many=True)
        return success_response(data=serializer.data) 