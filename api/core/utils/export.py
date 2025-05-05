"""
Utilities for API data export.

Provides mixins and functions to add export capabilities to API views.
"""

from rest_framework.response import Response
import csv, io, json


class ExportMixin:
    """
    Mixin to add export functionality to ViewSets and APIViews.
    
    Adds support for CSV, JSON, and Excel formats.
    """
    
    def is_export_request(self, request):
        """
        Check if the current request is for an export.
        
        Args:
            request: The request object
        
        Returns:
            bool: True if the request is for export, False otherwise
        """
        format_param = request.query_params.get('format', None)
        return format_param in ['csv', 'json', 'excel']
    
    def handle_export(self, request, queryset=None, serializer_class=None):
        """
        Handle export requests for different formats.
        
        Args:
            request: The request object
            queryset: The queryset to export (optional)
            serializer_class: Serializer class to use (optional)
            
        Returns:
            Response: The export response with appropriate content type
        """
        export_format = request.query_params.get('format', 'csv')
        
        # Use provided queryset or get it from the viewset
        if queryset is None:
            queryset = self.filter_queryset(self.get_queryset())
            
        # Use provided serializer class or get it from the viewset
        if serializer_class is None:
            serializer_class = self.get_serializer_class()
            
        # Serialize queryset
        serializer = serializer_class(queryset, many=True)
        data = serializer.data
        
        # Handle different export formats
        if export_format == 'csv':
            return self.export_as_csv(data, filename='export.csv')
        elif export_format == 'json':
            return self.export_as_json(data, filename='export.json')
        elif export_format == 'excel':
            return self.export_as_excel(data, filename='export.xlsx')
        else:
            # Format not supported
            return Response({
                'error': f'Export format not supported: {export_format}',
                'supported_formats': ['csv', 'json', 'excel']
            }, status=400)
    
    def export_as_csv(self, data, filename='export.csv'):
        """
        Export data as CSV.
        
        Args:
            data: The data to export
            filename: The filename for the export
            
        Returns:
            Response: CSV response
        """
        # If no data, return empty CSV with headers
        if not data:
            response = Response(
                content_type='text/csv',
                headers={'Content-Disposition': f'attachment; filename="{filename}"'}
            )
            response.content = ''
            return response
            
        # Get field names from the first item
        fieldnames = data[0].keys()
        
        # Create CSV
        csv_buffer = io.StringIO()
        writer = csv.DictWriter(csv_buffer, fieldnames=fieldnames)
        writer.writeheader()
        
        for item in data:
            # Convert non-string values to strings
            row = {k: str(v) if v is not None else '' for k, v in item.items()}
            writer.writerow(row)
            
        # Create response
        response = Response(
            csv_buffer.getvalue(),
            content_type='text/csv',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )
        
        return response
    
    def export_as_json(self, data, filename='export.json'):
        """
        Export data as JSON.
        
        Args:
            data: The data to export
            filename: The filename for the export
            
        Returns:
            Response: JSON response
        """
        response = Response(
            data,
            content_type='application/json',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )
        
        return response
    
    def export_as_excel(self, data, filename='export.xlsx'):
        """
        Export data as Excel.
        
        Args:
            data: The data to export
            filename: The filename for the export
            
        Returns:
            Response: Excel response
        """
        # If no data, return empty Excel with headers
        if not data:
            try:
                import openpyxl
                from openpyxl.utils import get_column_letter
                
                wb = openpyxl.Workbook()
                ws = wb.active
                
                # Add headers if we have field names
                if data:
                    for col_idx, header in enumerate(data[0].keys(), 1):
                        ws.cell(row=1, column=col_idx, value=header)
                
                # Save to buffer
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                
                # Create response
                response = Response(
                    buffer.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename="{filename}"'}
                )
                
                return response
            except ImportError:
                # Fallback to CSV if openpyxl is not available
                return self.export_as_csv(data, filename=filename.replace('.xlsx', '.csv'))
        
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Add headers
            for col_idx, header in enumerate(data[0].keys(), 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            # Add data
            for row_idx, item in enumerate(data, 2):
                for col_idx, (key, value) in enumerate(item.items(), 1):
                    ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else '')
            
            # Auto-size columns
            for col_idx, _ in enumerate(data[0].keys(), 1):
                ws.column_dimensions[get_column_letter(col_idx)].auto_size = True
            
            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Create response
            response = Response(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': f'attachment; filename="{filename}"'}
            )
            
            return response
        except ImportError:
            # Fallback to CSV if openpyxl is not available
            return self.export_as_csv(data, filename=filename.replace('.xlsx', '.csv')) 